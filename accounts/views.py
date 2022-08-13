from django.shortcuts import redirect, render
from django.http import HttpResponse, HttpResponseRedirect
from requests import request
from .forms import addUserProfileForm, levelForm
from django.contrib import messages
from .models import UserProfile, PassData, Levels, BlodsType
import qrcode, pytz, datetime, os, pdfkit
import os, pdfkit, shutil, zipfile, stat, errno
from pathlib import Path 

path = r"C:\Program Files\wkhtmltopdf\bin"
os.environ["PATH"] += os.pathsep + path

os.environ.update()

# Create your views here.

def index(request):
    if request.user.is_authenticated:
        return redirect('showUsers')
    else :
        return redirect('admin:index')

def addUser(request):
    if request.user.is_authenticated:
        userProfileForm = addUserProfileForm
        if request.method == "POST":
            userProfileForm = addUserProfileForm(request.POST)
            if userProfileForm.is_valid():
                userProfileForm.save()
                messages.error(request, "تم الاضافة بنجاح")
                return redirect('addUser')
            else:
                messages.error(request, "هناك خطاء في المدخلات")
                
        return render(request, 'account/addUser.html', {"form": userProfileForm})
    else:
        return redirect('index')

def QrCodeShow(request, id):
    if request.user.is_authenticated:
            
        userprofile = UserProfile.objects.get(id=id)
        
        
        # Encoding data using make() function
        img = qrcode.make(userprofile.pass_id)
        
        # Saving as an image file
        img.save('media/qr.png')
        
        return render(request, 'account/showQrCode.html', {"U":userprofile})
    else:
        return redirect('index')
        
def showUsers(request):
    if request.user.is_authenticated:
            
        search = ""
        try:
            search = request.GET['search']
        except:
            pass
        userprofile = UserProfile.objects.filter(username__contains=search)
        
        return render(request, 'account/usersShows.html', {"UA": userprofile})
    else:
        return redirect('index')

def EnterSchool(request):
    if request.user.is_authenticated:
            
        obj = {"sound":'nothing'}
        if request.method == 'POST':
            qrValue = request.POST['qrValue']
            userprofile = UserProfile.objects
            if userprofile.filter(pass_id=qrValue, is_Enabled=True).exists():
                
                obj["sound"] = 'true'
                pass_data = PassData.objects
                user = userprofile.get(pass_id=qrValue)
                if not user.GuardianNumber:
                    pass_data.create(user=user, username=user.username, school=user.school, GuardianNumber=0, is_sended=True, passDate=datetime.datetime.now(pytz.timezone('Asia/Baghdad'))).save()
                    messages.error(request, "لا يوجد رقم واتس لهاذا الطالب")
                    return redirect('enterSchool')
                pass_data.create(user=user, username=user.username, school=user.school, GuardianNumber=user.GuardianNumber, passDate=datetime.datetime.now(pytz.timezone('Asia/Baghdad'))).save()
            else:
                obj["sound"] = 'false'
            
        return render(request, 'account/EnterSchool.html', obj)
    else:
        return redirect('index')
def outSchool(request):
    return render(request, 'account/outSchool.html')


def UserViewer(request, id):
    if request.user.is_authenticated:
            
        userprofile = UserProfile.objects.get(id=id)
        return render(request, 'account/userViewer.html', {'U':userprofile})
    else:
        return redirect("index")

def getUserReport(request, id):
    if request.user.is_authenticated:
            
        userprofile = UserProfile.objects.filter(id=id)
        if userprofile.exists():
            userprofile = userprofile[0]
            if request.method == "POST":
                DateRang = request.POST['date-rang']
                DateRang = int(DateRang)
                FromDate = request.POST['dateFrom']
                FromDate = datetime.datetime.strptime(FromDate, "%Y-%m-%d")
                DateTo = request.POST['dateTo']
                DateTo = datetime.datetime.strptime(DateTo, "%Y-%m-%d")
                
                
                    
                start_date = FromDate
                end_date = DateTo    # perhaps date.now()

                delta = end_date - start_date   # returns timedelta
                listOfDate = []
                for i in range(delta.days + 1):
                    date = start_date + datetime.timedelta(days=i)
                    listOfDate.append(date)
                

                
                allRow = """
                """
                
                        
                rows = """ 
                    <tr class="row-start">
                    <td>{date}</td>
                    <td>{comeTime}</td>
                    <td>{outTime}</td>
                    </tr>
                    
                    """
                    
                htmlTemplate = open('template\PrintTemplate\getComeDaysData.html', 'r', encoding="UTF-8").read()
                
                base = datetime.datetime.today()
                #date_list = [base - datetime.timedelta(days=x) for x in range(DateRang)]
                date_list = listOfDate
                weekDaysAr = ["الاثنين", "الثلاثاء", "الاربعاء", "الخميس", "الجمعة", "السبت", "الاحد"]
                for i in date_list:
                    pass_data = PassData.objects.filter(passDate__date=i.date(), user__id=id)
                    #print( "from " + str(date_list[-1].date()) + " to " + str(datetime.datetime.now().date()))
                    if pass_data.exists():
                        start_pass = pass_data.filter(passDate__time__lte="11:00:00")
                        start_pass = start_pass.filter(passDate__time__gte="08:00:00")
                        
                        end_pass = pass_data.filter(passDate__time__gte="11:00:00")
                        end_pass = end_pass.filter(passDate__time__lte="16:00:00")
                        if start_pass:
                            start_pass = start_pass[0].passDate.time()
                        else:
                            start_pass = "لم ياتي"

                        if end_pass:
                            end_pass = end_pass[0].passDate.time()
                        else:
                            end_pass = "لم يخرج"
                            
                        #print(start_pass, end_pass)
                        
                        #print(i.date())
                        #print(weekDaysAr[i.weekday()])
                        #print("----------------------------------")

                        allRow += rows.format(date=str(i.date()) + "--" +  weekDaysAr[i.weekday()], comeTime=start_pass, outTime=end_pass)
                        
                #print(allRow)
                temolate_id = int(open('template/PrintTemplate/template_id/id.txt', 'r').read())
                htmlTemplate = htmlTemplate.format(dateFrom=str(date_list[0].date()), dateTo=str(date_list[-1].date()), Num=temolate_id, school=userprofile.school, username=userprofile.username, rows=allRow)
                    #pdfkit.from_file('template\PrintTemplate\getComeDaysData.html', 'passDataReport.pdf')
                htmlTemplate = htmlTemplate.replace("{{", "{")
                htmlTemplate = htmlTemplate.replace("}}", "}")
                pdfkit.from_string(htmlTemplate, 'media/template/passDataReport.pdf')
                open('template/PrintTemplate/template_id/id.txt', 'w').write(str(temolate_id + 1))
                return HttpResponseRedirect("/media/template/passDataReport.pdf")
                
        return render(request, 'account/getUserReport.html', {"id":id})

    else:
        return redirect('index')




BASE_DIR = Path(__file__).resolve().parent.parent
"""
def getAllUserReport(request):
    if request.user.is_authenticated:

        username_template = "" "




            <div style="vertical-align: top; position:relative; display: inline-block; width:45%; min-height:118px; background:none; " id="container_5d6e8fc5">
                <div style="margin: 10px; display: block; " id="container_5d6e8fc5_padding" >
                <div style="text-align:center;">
                    <div style="vertical-align: top; position:relative; display: inline-block; width:97%; min-height:42px; background:none; " id="container_4a8d779a">
                    <div style="margin: 10px; display: block; " id="container_4a8d779a_padding" >
                        <div style="text-align:center;">
                        <span style="font-size:14pt; font-family:Arial, Helvetica, sans-serif; color:#000000; font-weight:bold; ">اسم الطالب : {username}</span>
                        </div>
                        </div>
                    </div>
                    </div>
                <div style="clear:both"></div>
                </div>


                <table style="width:100%">
                <tr>
                    <td class="first-row">التاريخ</td>
                    <td class="first-row">وقت الحظور</td>
                    <td class="first-row">وقت الانصراف</td>
                </tr>

                {rows_template}
                
                </table>
                
                </div>




        " ""


        row_template = " ""


                <tr class="row-start">
                    <td>{date}</td>
                    <td>{comeTime}</td>
                    <td>{outTime}</td>
                    </tr>


        " ""    
        

        all_username_template = " ""
        "" "
        if request.method == "POST":
            shutil.rmtree(BASE_DIR / "media/template/allUserReport")
            os.mkdir(BASE_DIR / "media/template/allUserReport")
            
            DateRang = request.POST['date-rang']
            DateRang = int(DateRang)
            
            max_height_px = 1420
            px_now = 225
            username_px = 90
            evry_row_px = 30
            paper = 1
            column = 1
            
            base = datetime.datetime.today()
            date_list = [base - datetime.timedelta(days=x) for x in range(DateRang)]
            weekDaysAr = ["الاثنين", "الثلاثاء", "الاربعاء", "الخميس", "الجمعة", "السبت", "الاحد"]
            print( "from " + str(date_list[-1].date()) + " to " + str(datetime.datetime.now().date()))

            userprofile = UserProfile.objects.all()
            for user in userprofile:
                print("column = " + str(column))
                if column == 1:
                    px_now += username_px
                    px_now += evry_row_px * DateRang
                    if px_now > max_height_px:

                        last_template = open('template\PrintTemplate\getAllComeDaysData.html', 'r', encoding="UTF-8").read()
                        last_template = last_template.format(username_and_rows_template=all_username_template)
                        pdfkit.from_string(last_template, f'media/template/allUserReport/passDataReport{str(paper)}.pdf')
                        paper += 1
                        px_now = 225
                        all_username_template = """"""

                        px_now += username_px
                        px_now += evry_row_px * DateRang
                            
                        print("---------------------------------------------")
                    column = 2
                else:
                    column = 1
                    

                    
                #print(user.username)
                #print(px_now)
            
                template = username_template.format(username=user.username, rows_template="{rows_template}")
                all_rows = "" "
                "" "


                for i in date_list:

                    pass_data = PassData.objects.filter(passDate__date=i.date(), user__id=user.id)
                    
                    if pass_data.exists():
                        start_pass = pass_data.filter(passDate__time__lte="11:00:00")
                        start_pass = start_pass.filter(passDate__time__gte="08:00:00")
                        
                        end_pass = pass_data.filter(passDate__time__gte="11:00:00")
                        end_pass = end_pass.filter(passDate__time__lte="16:00:00")
                        if start_pass:
                            start_pass = start_pass[0].passDate.time()
                        else:
                            start_pass = "لم ياتي"

                        if end_pass:
                            end_pass = end_pass[0].passDate.time()
                        else:
                            end_pass = "لم يخرج"
                            
                        #print(start_pass, end_pass)
                        
                        #print(i.date())
                        #print(weekDaysAr[i.weekday()])
                        #print("----------------------------------")

                        all_rows += row_template.format(date=str(i.date()) + "--" +  weekDaysAr[i.weekday()], comeTime=start_pass, outTime=end_pass)
                all_username_template += "\n" + template.format(rows_template=all_rows)
            
            last_template = open('template\PrintTemplate\getAllComeDaysData.html', 'r', encoding="UTF-8").read()
            last_template = last_template.format(username_and_rows_template=all_username_template)
            
            pdfkit.from_string(last_template, f'media/template/allUserReport/passDataReport{str(paper)}.pdf')

                    
            def zipdir(path, ziph):
                # ziph is zipfile handle
                for root, dirs, files in os.walk(path):
                    for file in files:
                        ziph.write(os.path.join(root, file), 
                                os.path.relpath(os.path.join(root, file), 
                                                os.path.join(path, '..')))

            with zipfile.ZipFile('media/template/zip/AllUserReport.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipdir(BASE_DIR / "media/template/allUserReport", zipf)
            return HttpResponseRedirect("media/template/zip/AllUserReport.zip")
        
        return render(request, 'account/getAllUserReport.html', {"id":id})

    else:
        return redirect('index')


"""




def getAllUserReport(request):
    


    level_foem = levelForm()
    if request.user.is_authenticated:


        row_template = """

        <tr class="row-start">
          <td>{num}</td>
          <td>{username}</td>
          <td>{day}</td>
          <td>{date}</td>
          <td>{comeTime}</td>
          <td>{outTime}</td>
          <td>{note}</td>
        </tr>


        """    
        

        all_rows = """
        """
        if request.method == "POST":
            
            def handleRemoveReadonly(func, path, exc):
                excvalue = exc[1]
                if func in (os.rmdir, os.remove) and excvalue.errno == errno.EACCES:
                    os.chmod(path, stat.S_IRWXU| stat.S_IRWXG| stat.S_IRWXO) # 0777
                    func(path)
                else:
                    raise            
                

            shutil.rmtree(BASE_DIR / "media/template/allUserReport", ignore_errors=False, onerror=handleRemoveReadonly)
            os.mkdir(BASE_DIR / "media/template/allUserReport")
            
            DateRang = request.POST['date-rang']
            FromDate = request.POST['dateFrom']
            FromDate = datetime.datetime.strptime(FromDate, "%Y-%m-%d")
            DateTo = request.POST['dateTo']
            schoolPOST = request.POST['school']
            DateTo = datetime.datetime.strptime(DateTo, "%Y-%m-%d")
                    
            start_date = FromDate
            end_date = DateTo    # perhaps date.now()

            delta = end_date - start_date   # returns timedelta
            listOfDate = []
            for i in range(delta.days + 1):
                date = start_date + datetime.timedelta(days=i)
                listOfDate.append(date)
            
            level_id = request.POST['level']
            #print(level_id)
            DateRang = int(DateRang)
            
            max_height_px = 1420
            px_now = 225
            username_px = 90
            evry_row_px = 27
            paper = 1
            column = 1
            loop_counter = 0
            base = datetime.datetime.today()
            #date_list = [base - datetime.timedelta(days=x) for x in range(DateRang)]
            date_list = listOfDate
            weekDaysAr = ["الاثنين", "الثلاثاء", "الاربعاء", "الخميس", "الجمعة", "السبت", "الاحد"]
            #print( "from " + str(date_list[-1].date()) + " to " + str(datetime.datetime.now().date()))
            
            userprofile = UserProfile.objects.all().order_by('username').filter(level__id=level_id)
            if schoolPOST:
                if userprofile.filter(school=schoolPOST).exists():
                    userprofile = userprofile.filter(school=schoolPOST)
                else:
                    messages.info(request, "لا يوجد اي طالب في المدرسة الذي ادخلته")
                    return redirect('getAllUserReport')
                
            if not userprofile.exists():
                messages.error(request, "لا يوجد طلاب")
                return redirect('getAllUserReport')
            level_name = userprofile[0].level.name
            for user in userprofile:
                    
                    
                #print(user.username)
                #print(px_now)
            
                #template = username_template.format(username=user.username, rows_template="{rows_template}")
                #all_rows = """
                #"""


                for i in date_list:
                    print(px_now)
                    loop_counter += 1

                    #px_now += username_px
                    px_now += evry_row_px
                    if px_now > max_height_px:

                        last_template = open('template\PrintTemplate\getAllComeDaysData.html', 'r', encoding="UTF-8").read()
                        last_template = last_template.format(rows_template=all_rows, _from=str(date_list[0].date()), to=str(date_list[-1].date()), level=level_name)
                        pdfkit.from_string(last_template, f'media/template/allUserReport/passDataReport{str(paper)}.pdf')
                        paper += 1
                        px_now = 225
                        all_rows = """"""

                        #px_now += username_px
                        px_now += evry_row_px
                            
                        #print("---------------------------------------------")

                    pass_data = PassData.objects.filter(passDate__date=i.date(), user__id=user.id)
                    
                    if pass_data.exists():
                        start_pass = pass_data.filter(passDate__time__lte="11:00:00")
                        start_pass = start_pass.filter(passDate__time__gte="08:00:00")
                        
                        end_pass = pass_data.filter(passDate__time__gte="11:00:00")
                        end_pass = end_pass.filter(passDate__time__lte="16:00:00")
                        if start_pass:
                            start_pass = start_pass[0].passDate.time()
                        else:
                            start_pass = "لم ياتي"

                        if end_pass:
                            end_pass = end_pass[0].passDate.time()
                        else:
                            end_pass = "لم يخرج"
                            
                        #print(start_pass, end_pass)
                        
                        #print(i.date())
                        #print(weekDaysAr[i.weekday()])
                        #print("----------------------------------")
                    else:
                        end_pass = start_pass = "لم يأتي"
                    all_rows += row_template.format(date=str(i.date()), comeTime=start_pass, outTime=end_pass, num=loop_counter, username=user.username, day=weekDaysAr[i.weekday()], note=user.note)
                #all_username_template += "\n" + template.format(rows_template=all_rows)
            
            last_template = open('template\PrintTemplate\getAllComeDaysData.html', 'r', encoding="UTF-8").read()
            last_template = last_template.format(rows_template=all_rows, _from=str(date_list[0].date()), to=str(date_list[-1].date()), level=level_name)
            
            pdfkit.from_string(last_template, f'media/template/allUserReport/passDataReport{str(paper)}.pdf')

                    
            def zipdir(path, ziph):
                # ziph is zipfile handle
                for root, dirs, files in os.walk(path):
                    for file in files:
                        ziph.write(os.path.join(root, file), 
                                os.path.relpath(os.path.join(root, file), 
                                                os.path.join(path, '..')))

            with zipfile.ZipFile('media/template/zip/AllUserReport.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipdir(BASE_DIR / "media/template/allUserReport", zipf)
            return HttpResponseRedirect("media/template/zip/AllUserReport.zip")
        
        return render(request, 'account/getAllUserReport.html', {"id":id, "level_foem":level_foem})

    else:
        return redirect('index')



def BackupData(request):
    
    userprofile = UserProfile.objects.all()
    import xlsxwriter
    

    workbook = xlsxwriter.Workbook('media/backup/backup.xlsx')
    worksheet = workbook.add_worksheet()

    # Some data we want to write to the worksheet.
    expenses = [
    ]

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0
    worksheet.write(row, 0, 'اسم الطالب')
    worksheet.write(row, 1, 'المدرسة')
    worksheet.write(row, 2, 'ID المرحلة')
    worksheet.write(row, 3, 'المواليد')
    worksheet.write(row, 4, 'الرقم')
    worksheet.write(row, 5, 'الرقم 2')
    worksheet.write(row, 6, 'ID زمرة الدم')
    worksheet.write(row, 7, 'الامراض المزمنة')
    worksheet.write(row, 8, 'الملاحظات')
    worksheet.write(row, 9, 'qr ايدي المرور')
    worksheet.write(row, 10, 'هل هذا المستخدم مفعل')
    worksheet.write(row, 11, 'تاريخ الانشاء')
    worksheet.write(row, 12, 'ايدي المستخدم')
    row += 1
    for user in userprofile:
        #expenses + [user.username, user.school, user.level.id, user.born, user.GuardianNumber, user.importentNumber, user.blodType.id, user.diseases, user.note]
        #for username, school, level, born, GuardianNumber, importentNumber, blodType, diseases, note in (expenses):
        worksheet.write(row, col,     user.username)
        worksheet.write(row, col + 1, user.school)
        worksheet.write(row, col+ 2,     user.level.id)
        worksheet.write(row, col + 3, user.born)
        worksheet.write(row, col+ 4,     user.GuardianNumber)
        worksheet.write(row, col + 5, user.importentNumber)
        worksheet.write(row, col+ 6,     user.blodType.id)
        worksheet.write(row, col + 7, user.diseases)
        worksheet.write(row, col+ 8,     user.note)
        worksheet.write(row, col+ 9,     user.pass_id)
        worksheet.write(row, col+ 10,     user.is_Enabled)
        worksheet.write(row, col+ 11,     str(user.created_date))
        worksheet.write(row, col + 12,     user.id)
        col = 0
        row += 1
    # Start from the first cell. Rows and
    # Iterate over the data and write it out row by row.


    # Write a total using a formula.


    workbook.close()
    return HttpResponseRedirect('media/backup/backup.xlsx')


def importBackup(request):
    import openpyxl
    import random, string
    from django.core.files.storage import FileSystemStorage
    if request.method == 'POST':
        try:
            _file = request.FILES['file']
        except:
            messages.error(request, "يجب ادخل ملف النسخة الاحطياطية")
            return redirect('importBackup')
        #print(_file.name)
        fss = FileSystemStorage()
        try:
            os.remove('media/importBackup/backup.xlsx')
        except:
            pass
        _file = fss.save("importBackup/backup.xlsx", _file)
        loc = (BASE_DIR / "media/importBackup/backup.xlsx")


        wb = openpyxl.load_workbook(loc)
        ws = wb.active
        
                
        l = []

        l = [row for row in ws.iter_rows(max_col=13, values_only=True)]
        del l[0]
                
                
        def generate_random_password():
            characters = list(string.ascii_letters + string.digits + "!@#$%^&*()")
            ## length of password from the user
            length = 15

            ## shuffling the characters
            random.shuffle(characters)

            ## picking random characters from the list
            password = []
            for i in range(length):
                password.append(random.choice(characters))

            ## shuffling the resultant password
            random.shuffle(password)

            ## converting the list to string
            ## printing the list
            return "".join(password)

        userprofile = UserProfile.objects
        for username, school, level, born, GuardianNumber, importentNumber, blodType, diseases, note, pass_id, is_Enabled, created_date, userID in (l):
            #print(is_Enabled)
            level = Levels.objects.get(id=level)
            blodType = BlodsType.objects.get(id=blodType)
            if not userprofile.filter(id=userID).exists():
                if not pass_id:
                    pass_id = generate_random_password()
                if is_Enabled == "" or is_Enabled == " " or is_Enabled == 'TRUE' or is_Enabled=="None" or is_Enabled==None:
                    is_Enabled = True
                    
                if created_date:
                    created_date = created_date
                    
                else:
                    created_date = datetime.datetime.now()
                userprofile.create(username=username, school=school, level=level, born=born, GuardianNumber=GuardianNumber, importentNumber=importentNumber, blodType=blodType, diseases=diseases, note=note, pass_id=pass_id, is_Enabled=is_Enabled, created_date=created_date).save()
    return render(request, 'account/importBackup.html')
    